from openpyxl import load_workbook
import csv

# Excel-Datei und Blatt öffnen
wb = load_workbook("C:\\Users\\U6PGIT\\Desktop\\Übersichtsliste_Datensätze_test.xlsx")
sheet = wb.active  # oder wb["Sheetname"]

# CSV-Datei schreiben
with open("src\\sarah\\data\\csv\\ogd_datensaetze.csv", mode="w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f, delimiter=";")
    for row in sheet.iter_rows(values_only=True):
        writer.writerow(row)